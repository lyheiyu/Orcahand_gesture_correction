from __future__ import annotations

import argparse
import csv
import math
from collections import defaultdict
from pathlib import Path

import matplotlib
import numpy as np
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

matplotlib.use("Agg")
import matplotlib.pyplot as plt


CLASSIFIER_NAMES = {"svm": "SVM", "knn": "KNN", "rf": "RandomForest", "mlp": "MLP"}
REPRESENTATION_NAMES = {
    "joint_angle11": "JointAngle-11",
    "compact_optimized_action7": "Compact Refined-7",
}
SHEET_REPRESENTATION_NAMES = {
    "joint_angle11": "JointAngle",
    "compact_optimized_action7": "CompactRefined",
}
CMAPS = {"joint_angle11": "Blues", "compact_optimized_action7": "Oranges"}


def read_predictions(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def repeat_matrices(
    rows: list[dict[str, str]], labels: list[str]
) -> dict[tuple[str, str], np.ndarray]:
    label_index = {label: index for index, label in enumerate(labels)}
    grouped: dict[tuple[str, str, int], np.ndarray] = {}
    for row in rows:
        key = (row["classifier"], row["representation"], int(row["repeat"]))
        if key not in grouped:
            grouped[key] = np.zeros((len(labels), len(labels)), dtype=np.float64)
        grouped[key][label_index[row["true_label"]], label_index[row["predicted_label"]]] += 1.0

    normalized: dict[tuple[str, str], list[np.ndarray]] = defaultdict(list)
    for (classifier, representation, _), matrix in sorted(grouped.items()):
        totals = matrix.sum(axis=1, keepdims=True)
        values = np.divide(matrix, totals, out=np.zeros_like(matrix), where=totals != 0)
        normalized[(classifier, representation)].append(values)
    return {key: np.stack(values) for key, values in normalized.items()}


def statistics(stack: np.ndarray) -> dict[str, np.ndarray]:
    repeats = stack.shape[0]
    std = np.std(stack, axis=0, ddof=1) if repeats > 1 else np.zeros_like(stack[0])
    return {
        "mean": np.mean(stack, axis=0),
        "std": std,
        "ci95": 1.96 * std / math.sqrt(repeats),
        "min": np.min(stack, axis=0),
        "max": np.max(stack, axis=0),
    }


def style_matrix(ws, start_row: int, labels: list[str], title: str, values: np.ndarray) -> None:
    ws.cell(start_row, 1, title).font = Font(bold=True, size=12)
    header_row = start_row + 1
    ws.cell(header_row, 1, "True / Predicted").font = Font(bold=True)
    for column, label in enumerate(labels, start=2):
        cell = ws.cell(header_row, column, label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_index, label in enumerate(labels, start=header_row + 1):
        ws.cell(row_index, 1, label).font = Font(bold=True)
        for column_index, value in enumerate(values[row_index - header_row - 1], start=2):
            cell = ws.cell(row_index, column_index, float(value))
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")
    first = f"B{header_row + 1}"
    last = f"{get_column_letter(len(labels) + 1)}{header_row + len(labels)}"
    ws.conditional_formatting.add(
        f"{first}:{last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       end_type="num", end_value=1, end_color="2F75B5"),
    )


def plot_mean_std(
    mean: np.ndarray,
    std: np.ndarray,
    labels: list[str],
    title: str,
    cmap: str,
    output: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(7.5, 6.5))
    image = ax.imshow(mean, vmin=0.0, vmax=1.0, cmap=cmap)
    threshold = 0.52 * float(np.max(mean))
    for row in range(len(labels)):
        for column in range(len(labels)):
            color = "white" if mean[row, column] > threshold else "#20252B"
            ax.text(column, row, f"{mean[row, column]:.2f}\n±{std[row, column]:.2f}",
                    ha="center", va="center", fontsize=7.5, color=color)
    ax.set_xticks(np.arange(len(labels)), labels, rotation=35, ha="right", fontsize=8)
    ax.set_yticks(np.arange(len(labels)), labels, fontsize=8)
    ax.set_xlabel("Predicted label")
    ax.set_ylabel("True label")
    ax.set_title(title, weight="bold")
    fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04, label="Mean row-normalized frequency")
    fig.tight_layout()
    fig.savefig(output.with_suffix(".png"), dpi=300, bbox_inches="tight")
    fig.savefig(output.with_suffix(".pdf"), bbox_inches="tight")
    plt.close(fig)


def export_recall_stability(
    stacks: dict[tuple[str, str], np.ndarray],
    labels: list[str],
    output: Path,
    figure_dir: Path,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Per_Class_Recall"
    headers = [
        "classifier", "class",
        "jointangle_recall_mean", "jointangle_recall_std",
        "compact_recall_mean", "compact_recall_std",
        "compact_minus_joint_mean", "difference_std", "difference_ci95_half_width",
    ]
    summary.append(headers)
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    figure_dir.mkdir(parents=True, exist_ok=True)
    overview_fig, overview_axes = plt.subplots(2, 2, figsize=(15, 11), sharey=True)
    for classifier, overview_ax in zip(("svm", "knn", "rf", "mlp"), overview_axes.flat, strict=True):
        joint = np.diagonal(stacks[(classifier, "joint_angle11")], axis1=1, axis2=2)
        compact = np.diagonal(stacks[(classifier, "compact_optimized_action7")], axis1=1, axis2=2)
        classifier_rows = []
        for index, label in enumerate(labels):
            difference = compact[:, index] - joint[:, index]
            row = [
                CLASSIFIER_NAMES[classifier], label,
                float(np.mean(joint[:, index])), float(np.std(joint[:, index], ddof=1)),
                float(np.mean(compact[:, index])), float(np.std(compact[:, index], ddof=1)),
                float(np.mean(difference)), float(np.std(difference, ddof=1)),
                float(1.96 * np.std(difference, ddof=1) / math.sqrt(len(difference))),
            ]
            summary.append(row)
            classifier_rows.append(row)

        macro_joint = np.mean(joint, axis=1)
        macro_compact = np.mean(compact, axis=1)
        macro_difference = macro_compact - macro_joint
        macro_row = [
            CLASSIFIER_NAMES[classifier], "Macro average",
            float(np.mean(macro_joint)), float(np.std(macro_joint, ddof=1)),
            float(np.mean(macro_compact)), float(np.std(macro_compact, ddof=1)),
            float(np.mean(macro_difference)), float(np.std(macro_difference, ddof=1)),
            float(1.96 * np.std(macro_difference, ddof=1) / math.sqrt(len(macro_difference))),
        ]
        summary.append(macro_row)

        ws = workbook.create_sheet(CLASSIFIER_NAMES[classifier][:31])
        ws.append(headers[1:])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
        for row in classifier_rows + [macro_row]:
            ws.append(row[1:])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in range(2, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                ws.cell(row, column).number_format = "0.0000"
        for column in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(column)].width = 25

        x = np.arange(len(labels))
        width = 0.36
        joint_mean = np.mean(joint, axis=0)
        joint_std = np.std(joint, axis=0, ddof=1)
        compact_mean = np.mean(compact, axis=0)
        compact_std = np.std(compact, axis=0, ddof=1)

        def draw(ax: plt.Axes, title: str) -> None:
            ax.bar(x - width / 2, joint_mean, width, yerr=joint_std, capsize=3,
                   color="#4A7FA7", label="JointAngle-11")
            ax.bar(x + width / 2, compact_mean, width, yerr=compact_std, capsize=3,
                   color="#C8432F", label="Compact Refined-7")
            ax.set_xticks(x, labels, rotation=35, ha="right", fontsize=8)
            ax.set_ylim(0.25, 1.08)
            ax.set_ylabel("Per-class recall (mean ± std)")
            ax.set_title(title, weight="bold")
            ax.grid(axis="y", alpha=0.2)
            ax.legend(frameon=False, fontsize=8)

        fig, ax = plt.subplots(figsize=(9.5, 5.4))
        draw(ax, f"{CLASSIFIER_NAMES[classifier]} Per-class Recall Stability")
        fig.tight_layout()
        fig.savefig(figure_dir / f"recall_stability_{classifier}.png", dpi=300, bbox_inches="tight")
        fig.savefig(figure_dir / f"recall_stability_{classifier}.pdf", bbox_inches="tight")
        plt.close(fig)
        draw(overview_ax, CLASSIFIER_NAMES[classifier])

    summary.freeze_panes = "A2"
    summary.auto_filter.ref = summary.dimensions
    for column in range(3, summary.max_column + 1):
        for row in range(2, summary.max_row + 1):
            summary.cell(row, column).number_format = "0.0000"
    for column in range(1, summary.max_column + 1):
        summary.column_dimensions[get_column_letter(column)].width = 26

    overview_fig.suptitle("Per-class Recall Stability Across 20 Few-shot Training Selections", weight="bold")
    overview_fig.tight_layout(rect=(0, 0, 1, 0.97))
    overview_fig.savefig(figure_dir / "recall_stability_all_classifiers.png", dpi=300, bbox_inches="tight")
    overview_fig.savefig(figure_dir / "recall_stability_all_classifiers.pdf", bbox_inches="tight")
    plt.close(overview_fig)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def export(predictions_path: Path, workbook_path: Path, figure_dir: Path) -> None:
    rows = read_predictions(predictions_path)
    if not rows:
        raise ValueError("Prediction CSV is empty.")
    labels = sorted({row["true_label"] for row in rows} | {row["predicted_label"] for row in rows})
    stacks = repeat_matrices(rows, labels)
    stats = {key: statistics(stack) for key, stack in stacks.items()}
    figure_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "README"
    overview_rows = [
        ("Purpose", "Per-cell confusion-matrix stability across repeated few-shot training selections."),
        ("Input", str(predictions_path)),
        ("Repeats", len({int(row["repeat"]) for row in rows})),
        ("Classes", len(labels)),
        ("Normalization", "Each repeat is row-normalized before mean and sample std (ddof=1) are calculated."),
        ("Meaning of std", "Variation caused by different 3-shot training selections; not temporal trajectory stability."),
        ("CI", "95% CI half-width = 1.96 * sample_std / sqrt(number of repeats)."),
        ("Diagonal", "Per-class recall mean/std."),
        ("Off diagonal", "Mean/std of a specific misclassification route."),
    ]
    for row_index, (name, value) in enumerate(overview_rows, start=1):
        overview.cell(row_index, 1, name).font = Font(bold=True)
        overview.cell(row_index, 2, value)
    overview.column_dimensions["A"].width = 22
    overview.column_dimensions["B"].width = 105

    long_sheet = workbook.create_sheet("Cell_Statistics")
    headers = ["classifier", "representation", "true_label", "predicted_label",
               "mean", "sample_std", "ci95_half_width", "minimum", "maximum"]
    long_sheet.append(headers)
    for cell in long_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for (classifier, representation), values in sorted(stats.items()):
        for true_index, true_label in enumerate(labels):
            for predicted_index, predicted_label in enumerate(labels):
                long_sheet.append([
                    CLASSIFIER_NAMES[classifier], REPRESENTATION_NAMES[representation],
                    true_label, predicted_label,
                    float(values["mean"][true_index, predicted_index]),
                    float(values["std"][true_index, predicted_index]),
                    float(values["ci95"][true_index, predicted_index]),
                    float(values["min"][true_index, predicted_index]),
                    float(values["max"][true_index, predicted_index]),
                ])
    long_sheet.freeze_panes = "A2"
    long_sheet.auto_filter.ref = long_sheet.dimensions
    for column in range(5, 10):
        for cell in long_sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            cell[0].number_format = "0.0000"
    for column, width in enumerate((16, 22, 20, 20, 12, 12, 16, 12, 12), start=1):
        long_sheet.column_dimensions[get_column_letter(column)].width = width

    recall_sheet = workbook.create_sheet("Diagonal_Recall")
    recall_sheet.append(["classifier", "representation", "class", "recall_mean", "recall_std", "ci95_half_width"])
    for cell in recall_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
    summary_sheet = workbook.create_sheet("Method_Summary")
    summary_sheet.append(["classifier", "representation", "macro_recall_mean", "macro_recall_std",
                          "mean_cell_std", "max_cell_std"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")

    for (classifier, representation), values in sorted(stats.items()):
        diagonal_by_repeat = np.diagonal(stacks[(classifier, representation)], axis1=1, axis2=2)
        macro_recall_by_repeat = np.mean(diagonal_by_repeat, axis=1)
        for index, label in enumerate(labels):
            recall_sheet.append([
                CLASSIFIER_NAMES[classifier], REPRESENTATION_NAMES[representation], label,
                float(values["mean"][index, index]), float(values["std"][index, index]),
                float(values["ci95"][index, index]),
            ])
        summary_sheet.append([
            CLASSIFIER_NAMES[classifier], REPRESENTATION_NAMES[representation],
            float(np.mean(macro_recall_by_repeat)),
            float(np.std(macro_recall_by_repeat, ddof=1)),
            float(np.mean(values["std"])),
            float(np.max(values["std"])),
        ])

        sheet_name = f"{CLASSIFIER_NAMES[classifier][:8]}_{SHEET_REPRESENTATION_NAMES[representation]}"
        ws = workbook.create_sheet(sheet_name[:31])
        ws.freeze_panes = "B3"
        style_matrix(ws, 1, labels, "Mean row-normalized confusion matrix", values["mean"])
        style_matrix(ws, 11, labels, "Sample standard deviation across repeats", values["std"])
        style_matrix(ws, 21, labels, "95% confidence-interval half-width", values["ci95"])
        style_matrix(ws, 31, labels, "Minimum across repeats", values["min"])
        style_matrix(ws, 41, labels, "Maximum across repeats", values["max"])
        ws.column_dimensions["A"].width = 23
        for column in range(2, len(labels) + 2):
            ws.column_dimensions[get_column_letter(column)].width = 18

        plot_mean_std(
            values["mean"], values["std"], labels,
            f"{CLASSIFIER_NAMES[classifier]} - {REPRESENTATION_NAMES[representation]}\nMean ± std across repeated training selections",
            CMAPS[representation], figure_dir / f"cm_stability_{classifier}_{representation}",
        )

    for ws in (recall_sheet, summary_sheet):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in range(4, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                ws.cell(row, column).number_format = "0.0000"
        for column in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(column)].width = 22

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    export_recall_stability(
        stacks,
        labels,
        workbook_path.with_name("per_class_recall_stability.xlsx"),
        figure_dir.parent / "recall_stability",
    )
    print(f"workbook={workbook_path}")
    print(f"recall_workbook={workbook_path.with_name('per_class_recall_stability.xlsx')}")
    print(f"stability_figures={len(stacks)} labels={len(labels)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export per-cell confusion-matrix mean/std to Excel.")
    parser.add_argument(
        "--predictions",
        default="paper_final_compact_orca_20260827/figures/supplementary_confusion_matrices/frozen_final_predictions.csv",
    )
    parser.add_argument(
        "--output",
        default="paper_final_compact_orca_20260827/figures/supplementary_confusion_matrices/cm_stability_statistics.xlsx",
    )
    parser.add_argument(
        "--figure-dir",
        default="paper_final_compact_orca_20260827/figures/supplementary_confusion_matrices/stability_mean_std",
    )
    args = parser.parse_args()
    export(Path(args.predictions).resolve(), Path(args.output).resolve(), Path(args.figure_dir).resolve())


if __name__ == "__main__":
    main()
